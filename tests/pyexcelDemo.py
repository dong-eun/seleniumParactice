import openpyxl

testDic = {}

workBook = openpyxl.load_workbook('../testData/testData.xlsx')
sheet = workBook.active
for i in range(2, sheet.max_row+1):
  if sheet.cell(row=i, column=1).value == "TestCase2":
    for j in range(2, sheet.max_column+1):
      testDic[sheet.cell(row=1, column=j).value] = sheet.cell(row=i,column=j).value

