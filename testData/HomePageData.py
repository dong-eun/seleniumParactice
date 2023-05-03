import openpyxl

class HomePageData:

  test_HomePage_data = [
    {
      "firstname":"Dominic",
      "email":"dominic@page.com",
      "gender": "Male"
    },
    {
      "firstname": "Dominic2",
      "email": "dominic2@page.com",
      "gender": "Female"
    },
    {
      "firstname": "Dominic3",
      "email": "dominic3@page.com",
      "gender": "Male"
    }
  ]

  @staticmethod
  def getTestData(test_case_name):
    testDic = {}
    workBook = openpyxl.load_workbook(
      './testData.xlsx')
    sheet = workBook.active

    for i in range(2, sheet.max_row + 1):
      if sheet.cell(row=i, column=1).value == test_case_name:
        for j in range(2, sheet.max_column + 1):
          testDic[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

    return [testDic]